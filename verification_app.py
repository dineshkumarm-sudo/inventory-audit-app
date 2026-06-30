%%writefile verification_app.py
import streamlit as st
import pandas as pd
import re
import openpyxl
from openpyxl.styles import PatternFill
import io

st.set_page_config(page_title="Data Audit & Rounds Verifier", layout="centered")
st.title("🛡️ Inventory Data Audit Engine")
st.write("Upload your product export spreadsheet to run the metadata, bullet type, and round-count audit engine.")

uploaded_file = st.file_uploader("Choose your Excel file (.xlsx)", type=["xlsx"])

# Broad lookup dictionary of common ammunition projectile/bullet types
BULLET_TYPES_VOCAB = [
    "FMJ", "CPHP", "JHP", "HP", "SP", "LRN", "TMJ", "OTM", "FMC", "BT", 
    "SCHP", "SJHP", "JSP", "LHP", "FMJBT", "BTHP", "FTX", "VMAX", "V-MAX",
    "MC", "AP", "Subsonic"
]

if uploaded_file is not None:
    df = pd.read_excel(uploaded_file)
    st.success(f"Successfully loaded {len(df)} rows!")
    
    column_options = df.columns.tolist()
    
    col1, col2 = st.columns(2)
    with col1:
        title_col = st.selectbox("Select Title Column:", column_options, 
                                 index=column_options.index('Generated Title') if 'Generated Title' in column_options else 0)
    with col2:
        desc_col = st.selectbox("Select Description Column:", column_options, 
                                index=column_options.index('Generated Description') if 'Generated Description' in column_options else 0)

    if st.button("🚀 Run Complete Dual-Engine Audit"):
        for col in ['MPN', 'UPC', 'Brand']:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
                df[col] = df[col].replace('nan', '')

        desc_position = df.columns.get_loc(desc_col)
        
        new_cols = ['Extracted Brand', 'Extracted UPC', 'Extracted MPN', 'Extracted Bullet Type', 'Extracted Title Rounds', 
                    'Extracted Desc Quantity', 'Extracted Packaging Rounds', 'Metadata Verification Comment', 'Rounds Verification Comment']
        for c in new_cols:
            if c in df.columns:
                df = df.drop(columns=[c])
                
        df.insert(desc_position + 1, 'Extracted Brand', None)
        df.insert(desc_position + 2, 'Extracted UPC', None)
        df.insert(desc_position + 3, 'Extracted MPN', None)
        df.insert(desc_position + 4, 'Extracted Bullet Type', None)
        df.insert(desc_position + 5, 'Extracted Title Rounds', None)
        df.insert(desc_position + 6, 'Extracted Desc Quantity', None)
        df.insert(desc_position + 7, 'Extracted Packaging Rounds', None)
        df.insert(desc_position + 8, 'Metadata Verification Comment', None)
        df.insert(desc_position + 9, 'Rounds Verification Comment', None)

        excel_buffer = io.BytesIO()
        df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)
        
        wb = openpyxl.load_workbook(excel_buffer)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        orig_brand_idx = headers.index('Brand') + 1 if 'Brand' in headers else 1
        orig_mpn_idx = headers.index('MPN') + 1 if 'MPN' in headers else 1
        orig_upc_idx = headers.index('UPC') + 1 if 'UPC' in headers else 1

        title_col_idx = headers.index(title_col) + 1
        desc_col_idx = headers.index(desc_col) + 1

        ext_brand_idx = headers.index('Extracted Brand') + 1
        ext_upc_idx = headers.index('Extracted UPC') + 1
        ext_mpn_idx = headers.index('Extracted MPN') + 1
        ext_bullet_idx = headers.index('Extracted Bullet Type') + 1

        ext_t_rounds_idx = headers.index('Extracted Title Rounds') + 1
        ext_d_qty_idx = headers.index('Extracted Desc Quantity') + 1
        ext_p_rounds_idx = headers.index('Extracted Packaging Rounds') + 1

        meta_comment_idx = headers.index('Metadata Verification Comment') + 1
        rounds_comment_idx = headers.index('Rounds Verification Comment') + 1

        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        progress_bar = st.progress(0)
        status_text = st.empty()

        total_rows = ws.max_row
        for row_idx in range(2, total_rows + 1):
            status_text.text(f"Auditing Row {row_idx-1} of {total_rows-1}...")
            
            orig_brand = str(ws.cell(row=row_idx, column=orig_brand_idx).value).strip() if ws.cell(row=row_idx, column=orig_brand_idx).value is not None else ""
            orig_mpn = str(ws.cell(row=row_idx, column=orig_mpn_idx).value).strip() if ws.cell(row=row_idx, column=orig_mpn_idx).value is not None else ""
            orig_upc = str(ws.cell(row=row_idx, column=orig_upc_idx).value).strip() if ws.cell(row=row_idx, column=orig_upc_idx).value is not None else ""
            
            title_text = str(ws.cell(row=row_idx, column=title_col_idx).value) if ws.cell(row=row_idx, column=title_col_idx).value else ""
            desc_cell = ws.cell(row=row_idx, column=desc_col_idx)
            desc_text = str(desc_cell.value) if desc_cell.value else ""
            
            # Metadata patterns parsing
            brand_match = re.search(r'Brand:\s*<\/strong>\s*([^<]+)', desc_text, re.IGNORECASE)
            if not brand_match:
                brand_match = re.search(r'Brand:.*?<\/strong>\s*([^<]+)', desc_text, re.IGNORECASE)
                
            upc_match = re.search(r'UPC:\s*<\/strong>\s*([0-9a-zA-Z]+)', desc_text, re.IGNORECASE)
            if not upc_match:
                upc_match = re.search(r'UPC:.*?<\/strong>\s*([0-9a-zA-Z]+)', desc_text, re.IGNORECASE)
            if not upc_match:
                upc_match = re.search(r'UPC:\s*([0-9a-zA-Z]+)', desc_text, re.IGNORECASE)
                
            mpn_match = re.search(r'MPN:\s*<\/strong>\s*([0-9a-zA-Z-]+)', desc_text, re.IGNORECASE)
            if not mpn_match:
                mpn_match = re.search(r'MPN:.*?<\/strong>\s*([0-9a-zA-Z-]+)', desc_text, re.IGNORECASE)
            if not mpn_match:
                mpn_match = re.search(r'MPN:\s*([0-9a-zA-Z-]+)', desc_text, re.IGNORECASE)
                
            # Precision Bullet Acronym Extraction Logic
            extracted_bullet = ""
            title_prefix = title_text.split("-")[0].strip()
            
            # Method 1: Check vocabulary entries explicitly
            found_bullets = [b for b in BULLET_TYPES_VOCAB if re.search(rf'\b{re.escape(b)}\b', title_prefix, re.IGNORECASE)]
            if found_bullets:
                # Pick the match closest to the end of the prefix, as bullet types follow calibers
                extracted_bullet = max(found_bullets, key=lambda b: title_prefix.upper().find(b.upper()))
            else:
                # Method 2 Fallback: If not in vocab dictionary, extract remaining standard acronym blocks
                bullet_matches = re.findall(r'\b([A-Z]{2,4})\b', title_prefix)
                bullet_filtered = [b for b in bullet_matches if b != orig_brand.upper() and b not in ["LR", "ACPC", "GA"]]
                if bullet_filtered:
                    extracted_bullet = bullet_filtered[-1]
            
            extracted_brand = brand_match.group(1).strip() if brand_match else ""
            extracted_upc = upc_match.group(1).strip() if upc_match else ""
            extracted_mpn = mpn_match.group(1).strip() if mpn_match else ""
            
            if extracted_brand: ws.cell(row=row_idx, column=ext_brand_idx).value = str(extracted_brand)
            if extracted_bullet: ws.cell(row=row_idx, column=ext_bullet_idx).value = str(extracted_bullet).upper()
            if extracted_upc:
                ws.cell(row=row_idx, column=ext_upc_idx).value = str(extracted_upc)
                ws.cell(row=row_idx, column=ext_upc_idx).number_format = '@'
            if extracted_mpn:
                ws.cell(row=row_idx, column=ext_mpn_idx).value = str(extracted_mpn)
                ws.cell(row=row_idx, column=ext_mpn_idx).number_format = '@'

            # Parse quantities & round counts
            t_rounds_match = re.search(r'-\s*(\d+)\s*Rounds', title_text, re.IGNORECASE)
            if not t_rounds_match:
                t_rounds_match = re.search(r'(\d+)\s*Rounds', title_text, re.IGNORECASE)
                
            d_qty_match = re.search(r'Quantity:.*?<\/strong>\s*(\d+)\s*Rounds', desc_text, re.IGNORECASE)
            if not d_qty_match:
                d_qty_match = re.search(r'Quantity:\s*(\d+)\s*Rounds', desc_text, re.IGNORECASE)
                
            p_rounds_match = re.search(r'Packaging Information.*?<\/h3>\s*<p>\s*(\d+)\s*Rounds', desc_text, re.IGNORECASE | re.DOTALL)
            if not p_rounds_match:
                p_rounds_match = re.search(r'Packaging Information.*?\s*(\d+)\s*Rounds', desc_text, re.IGNORECASE | re.DOTALL)
                
            val_t_rounds = t_rounds_match.group(1).strip() if t_rounds_match else ""
            val_d_qty = d_qty_match.group(1).strip() if d_qty_match else ""
            val_p_rounds = p_rounds_match.group(1).strip() if p_rounds_match else ""
            
            if val_t_rounds: ws.cell(row=row_idx, column=ext_t_rounds_idx).value = int(val_t_rounds)
            if val_d_qty: ws.cell(row=row_idx, column=ext_d_qty_idx).value = int(val_d_qty)
            if val_p_rounds: ws.cell(row=row_idx, column=ext_p_rounds_idx).value = int(val_p_rounds)

            # Metadata Engine cross check evaluations
            meta_comments = []
            if 'Brand' in headers and extracted_brand and orig_brand:
                if extracted_brand.lower() != orig_brand.lower(): meta_comments.append(f"Brand Mismatch ({orig_brand} vs {extracted_brand})")
            elif extracted_brand: meta_comments.append("Brand added from Desc")
                
            if 'UPC' in headers and extracted_upc and orig_upc:
                if extracted_upc != orig_upc: meta_comments.append(f"UPC Mismatch ({orig_upc} vs {extracted_upc})")
            elif extracted_upc: meta_comments.append("UPC added from Desc")
                
            if 'MPN' in headers and extracted_mpn and orig_mpn:
                if extracted_mpn.lower() != orig_mpn.lower(): meta_comments.append(f"MPN Mismatch ({orig_mpn} vs {extracted_mpn})")
            elif extracted_mpn: meta_comments.append("MPN added from Desc")
                
            # Bullet validation cross check string evaluation
            if extracted_bullet:
                if extracted_bullet.upper() in desc_text.upper():
                    meta_comments.append(f"Bullet Type ({extracted_bullet.upper()}): Validated")
                else:
                    meta_comments.append(f"Bullet Type Mismatch: '{extracted_bullet.upper()}' not found in description")
                
            if not extracted_upc and "UPC" in desc_text.upper():
                desc_cell.fill = red_fill
                meta_comments.append("UPC Syntax Alert: Pattern broke extraction execution")
                
            if meta_comments:
                ws.cell(row=row_idx, column=meta_comment_idx).value = " | ".join(meta_comments)
            else:
                ws.cell(row=row_idx, column=meta_comment_idx).value = "All Metadata Validated"

            # Rounds Engine cross check evaluations
            rounds_map = {"Title": val_t_rounds, "Quantity": val_d_qty, "Packaging": val_p_rounds}
            missing_rounds = [k for k, v in rounds_map.items() if not v]
            
            if missing_rounds:
                present_data = ", ".join([f"{k}({v})" for k, v in rounds_map.items() if v])
                ws.cell(row=row_idx, column=rounds_comment_idx).value = f"Data Insufficient: Missing {', '.join(missing_rounds)} | Found: {present_data}"
            elif val_t_rounds == val_d_qty == val_p_rounds:
                ws.cell(row=row_idx, column=rounds_comment_idx).value = "Rounds Match"
            else:
                ws.cell(row=row_idx, column=rounds_comment_idx).value = f"Mismatch: Title ({val_t_rounds}) vs Quantity ({val_d_qty}) vs Packaging ({val_p_rounds})"

            progress_bar.progress((row_idx - 1) / (total_rows - 1))
            
        status_text.text("🎉 Processing and Excel data synthesis complete!")
        
        out_buffer = io.BytesIO()
        wb.save(out_buffer)
        out_buffer.seek(0)
        
        st.download_button(
            label="📥 Download Audited Excel File",
            data=out_buffer.getvalue(),
            file_name="Audited_Product_Data.xlsx",
            mime="application/vnd.ms-excel"
        )
